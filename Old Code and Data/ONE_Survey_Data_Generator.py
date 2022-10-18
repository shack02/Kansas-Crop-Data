# ----------------------------------------------------------------------------------------------------------------------
# Name: ONE_Survey_Data_Generator.py
# Created on: 3/15/2022
# Created by: Sean Hackenberg
# Description: This script contains functions used to extract survey data from a spreadsheet and plot this data on a
# graph
# Input:
#  - SURVEY Acres Planted.xlsx or SURVEY Acres Harvested.xlsx
# Output:
#  - List of years with data for each county
#  - List of values for each year for each county
#  - List of counties
#  - Graph of survey data per year for each county with data
# ----------------------------------------------------------------------------------------------------------------------
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


# function used to extract survey data of cotton harvested or planted from Survey spreadsheets
def extract_survey_data():
    wb = Workbook()
    counties = []
    sheets = []
    county_time_line = []
    county_yearly_value = []
    # change filename to correct path with format r"filepath"
    filename = r"SURVEY Acres Planted.xlsx"
    # filename = r"SURVEY Acres Planted.xlsx"
    wb = load_workbook(filename)  # loads file into the workbook
    ws = wb.active  # sets up the workbook

    for i, s in enumerate(wb.sheetnames):
        counties.append(s)  # adds the name of each county's sheet to the counties list

    del counties[-1]

    for county in counties:  # goes through each sheet in file and
        ws = wb[county]  # sets the active page to the name of the current county
        column_t = ws['T']  # creates a list of the data in column t
        list = []
        for cell in column_t:  # gets cell data from column T and appends list to county_yearly_value list
            if cell.value != "Value":
                list.append(cell.value)
        county_yearly_value.append(list)
        column_b = ws['B']  # creates a list of the data in column b
        list = []
        for cell in column_b:  # gets cell data from column B and appends list to county_time_line list
            if cell.value != "Year":
                if cell.value in list:
                    county_yearly_value[counties.index(county)][list.index(cell.value) + 1] = \
                        county_yearly_value[counties.index(county)][list.index(cell.value)] + \
                        county_yearly_value[counties.index(county)][list.index(cell.value) + 1]
                    del county_yearly_value[counties.index(county)][list.index(cell.value)]
                if cell.value not in list:
                    list.append(cell.value)
        county_time_line.append(list)
    return county_time_line, county_yearly_value, counties


# function used to plot graphs of the data extracted with survey_data
def plot_survey_graphs(county_time_line, county_yearly_value, counties):
    for county in counties:  # plots the graph for each county on separate pages
        current_index = counties.index(county)
        x_axis = county_time_line[current_index]
        y_axis = county_yearly_value[current_index]
        plt.bar(x_axis, y_axis)
        plt.xticks(x_axis)
        plt.title(county + " Survey: Cotton Upland, Acres Planted")
        plt.xlabel("Year")
        plt.ylabel("Value")
        # plot = plt.figure(counties.index(county))
        plt.show()

